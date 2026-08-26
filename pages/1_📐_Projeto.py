import streamlit as st
import pandas as pd
from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image as XLImage


# ============================================================
# CONFIGURAÇÃO DA PÁGINA
# ============================================================

st.set_page_config(
    page_title="Calculadora Steel Framing",
    page_icon="📐",
    layout="wide",
)


# ============================================================
# CSS
# ============================================================

st.markdown(
    """
    <style>

    html,
    body,
    [data-testid="stAppViewContainer"],
    .stApp {
        font-family: "Inter", "Segoe UI", Roboto, Helvetica, Arial, sans-serif !important;
    }

    .stApp {
        background: #f5f7fa !important;
    }

    .block-container {
        max-width: 1400px;
        padding-top: 2rem;
        padding-bottom: 4rem;
    }

    div[data-testid="stVerticalBlock"] > div:has(h1) {
        background: linear-gradient(
            135deg,
            #17202a 0%,
            #263746 55%,
            #34495e 100%
        ) !important;

        border-radius: 18px;
        padding: 35px 38px 30px 38px;
        margin-bottom: 25px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.12);
    }

    div[data-testid="stVerticalBlock"] h1 {
        color: #6fa8c9 !important;
        -webkit-text-fill-color: #6fa8c9 !important;
        font-size: 3.5rem !important;
        line-height: 1.2 !important;
        font-weight: 900 !important;
        letter-spacing: -1px !important;
        margin: 0 0 10px 0 !important;
        border: none !important;
    }

    div[data-testid="stVerticalBlock"] h1 + p {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        font-size: 1.2rem !important;
        line-height: 1.6 !important;
        font-weight: 500 !important;
    }

    h1,
    h2,
    h3,
    h4,
    h5,
    h6 {
        color: #17202a !important;
        opacity: 1 !important;
        visibility: visible !important;
    }

    p,
    span,
    div {
        opacity: 1;
    }

    label {
        color: #17202a !important;
        opacity: 1 !important;
        visibility: visible !important;
        -webkit-text-fill-color: #17202a !important;
    }

    label p {
        color: #17202a !important;
        font-weight: 600 !important;
        opacity: 1 !important;
        visibility: visible !important;
        -webkit-text-fill-color: #17202a !important;
    }

    [data-testid="stMarkdownContainer"] {
        color: #17202a !important;
        opacity: 1 !important;
        visibility: visible !important;
    }

    [data-testid="stMarkdownContainer"] p {
        color: #17202a !important;
        opacity: 1 !important;
        visibility: visible !important;
    }

    [data-testid="stMarkdownContainer"] strong {
        color: #17202a !important;
        opacity: 1 !important;
        visibility: visible !important;
    }

    div[data-testid="stTextInput"] label,
    div[data-testid="stNumberInput"] label,
    div[data-testid="stDateInput"] label,
    div[data-testid="stSelectbox"] label,
    div[data-testid="stTextArea"] label,
    div[data-testid="stRadio"] label,
    div[data-testid="stCheckbox"] label {

        color: #17202a !important;
        -webkit-text-fill-color: #17202a !important;
        opacity: 1 !important;
        visibility: visible !important;
        display: block !important;
        font-weight: 600 !important;
    }

    div[data-testid="stTextInput"] label p,
    div[data-testid="stNumberInput"] label p,
    div[data-testid="stDateInput"] label p,
    div[data-testid="stSelectbox"] label p,
    div[data-testid="stTextArea"] label p,
    div[data-testid="stRadio"] label p,
    div[data-testid="stCheckbox"] label p {

        color: #17202a !important;
        -webkit-text-fill-color: #17202a !important;
        opacity: 1 !important;
        visibility: visible !important;
        display: block !important;
        font-weight: 600 !important;
    }

    input,
    textarea,
    select {
        color: #17202a !important;
        -webkit-text-fill-color: #17202a !important;
        background-color: #ffffff !important;
        opacity: 1 !important;
        visibility: visible !important;
    }

    input::placeholder,
    textarea::placeholder {
        color: #7f8c8d !important;
        -webkit-text-fill-color: #7f8c8d !important;
        opacity: 1 !important;
    }

    input:focus,
    textarea:focus,
    select:focus {
        color: #17202a !important;
        -webkit-text-fill-color: #17202a !important;
    }

    div[data-testid="stNumberInput"] {
        background: #ffffff !important;
        border-radius: 10px !important;
    }

    div[data-testid="stNumberInput"] input {
        color: #17202a !important;
        background-color: #ffffff !important;
        caret-color: #17202a !important;
        -webkit-text-fill-color: #17202a !important;
        opacity: 1 !important;
        visibility: visible !important;
        font-weight: 500 !important;
    }

    div[data-testid="stNumberInput"] button {
        color: #17202a !important;
        background-color: #ffffff !important;
    }

    div[data-testid="stNumberInput"] button:hover {
        color: #17202a !important;
        background-color: #e9eef2 !important;
    }

    div[data-testid="stTextInput"] {
        background: #ffffff !important;
    }

    div[data-testid="stTextInput"] input {
        color: #17202a !important;
        background-color: #ffffff !important;
        caret-color: #17202a !important;
        -webkit-text-fill-color: #17202a !important;
        opacity: 1 !important;
        visibility: visible !important;
    }

    div[data-testid="stDateInput"] input {
        color: #17202a !important;
        background-color: #ffffff !important;
        -webkit-text-fill-color: #17202a !important;
        opacity: 1 !important;
        visibility: visible !important;
    }

    div[data-baseweb="input"] {
        background-color: #ffffff !important;
    }

    div[data-baseweb="input"] input {
        color: #17202a !important;
        -webkit-text-fill-color: #17202a !important;
        background-color: #ffffff !important;
        opacity: 1 !important;
        visibility: visible !important;
        caret-color: #17202a !important;
        font-weight: 500 !important;
    }

    div[data-baseweb="select"] {
        background-color: #ffffff !important;
        color: #17202a !important;
    }

    div[data-baseweb="select"] * {
        color: #17202a !important;
        -webkit-text-fill-color: #17202a !important;
        opacity: 1 !important;
    }

    div[data-testid="stRadio"] label,
    div[data-testid="stRadio"] label p,
    div[data-testid="stCheckbox"] label,
    div[data-testid="stCheckbox"] label p {
        color: #17202a !important;
        -webkit-text-fill-color: #17202a !important;
        opacity: 1 !important;
        visibility: visible !important;
    }

    div[data-testid="stTextArea"] textarea {
        color: #17202a !important;
        -webkit-text-fill-color: #17202a !important;
        background-color: #ffffff !important;
        opacity: 1 !important;
    }

    .nome-material {
        color: #17202A !important;
        background-color: #ffffff !important;
        font-size: 1.25rem !important;
        font-weight: 800 !important;
        margin: 10px 0 12px 0 !important;
        padding: 8px 12px !important;
        display: block !important;
        visibility: visible !important;
        opacity: 1 !important;
        border-radius: 8px !important;
        border-left: 5px solid #6fa8c9 !important;
    }

    .stWrite {
        color: #17202a !important;
    }

    div.stButton > button {
        color: #17202a !important;
        background-color: #ffffff !important;
        border: 1px solid #b7c3cc !important;
        font-weight: 700 !important;
        border-radius: 10px !important;
    }

    div.stButton > button:hover {
        color: #17202a !important;
        background-color: #e9eef2 !important;
        border-color: #6fa8c9 !important;
    }

    div.stButton > button[kind="primary"] {
        height: 60px;
        color: #ffffff !important;
        background: linear-gradient(
            135deg,
            #34495e,
            #17202a
        ) !important;

        font-size: 1.25rem;
        font-weight: 800;
        border-radius: 12px;
        width: 100%;
        border: none !important;
    }

    div.stButton > button[kind="primary"] p,
    div.stButton > button[kind="primary"] span {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
    }

    div.stButton > button[kind="primary"]:hover {
        color: #ffffff !important;
        background: linear-gradient(
            135deg,
            #45627a,
            #263746
        ) !important;
    }

    div[data-testid="stDownloadButton"] button {
        color: #ffffff !important;
        background: linear-gradient(
            135deg,
            #217346,
            #185c37
        ) !important;

        border: none !important;
        min-height: 55px;
        font-size: 1.1rem !important;
        font-weight: 800 !important;
        border-radius: 10px !important;
    }

    div[data-testid="stDownloadButton"] button p,
    div[data-testid="stDownloadButton"] button span {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
    }

    div[data-testid="stDownloadButton"] button:hover {
        color: #ffffff !important;
        background: linear-gradient(
            135deg,
            #2e8b57,
            #217346
        ) !important;
    }

    details,
    [data-testid="stExpander"] {
        background-color: #ffffff !important;
        border-radius: 10px !important;
    }

    [data-testid="stExpander"] summary {
        color: #17202a !important;
        background-color: #ffffff !important;
        font-weight: 700 !important;
    }

    [data-testid="stExpander"] summary p {
        color: #17202a !important;
        font-weight: 700 !important;
    }

    [data-testid="stExpander"] summary span {
        color: #17202a !important;
    }

    button[data-baseweb="tab"] {
        color: #17202a !important;
        font-weight: 700 !important;
    }

    button[data-baseweb="tab"][aria-selected="true"] {
        color: #17202a !important;
        font-weight: 800 !important;
    }

    section[data-testid="stSidebar"] {
        background-color: #ffffff !important;
    }

    section[data-testid="stSidebar"] * {
        color: #17202a !important;
    }

    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] label p {
        color: #17202a !important;
        -webkit-text-fill-color: #17202a !important;
        opacity: 1 !important;
    }

    section[data-testid="stSidebar"] input {
        color: #17202a !important;
        background-color: #ffffff !important;
        -webkit-text-fill-color: #17202a !important;
    }

    [data-testid="stMetricLabel"],
    [data-testid="stMetricLabel"] p,
    [data-testid="stMetricValue"],
    [data-testid="stMetricDelta"] {
        color: #17202a !important;
        -webkit-text-fill-color: #17202a !important;
    }

    [data-testid="stAlert"],
    [data-testid="stAlert"] p {
        color: #17202a !important;
    }

    [data-testid="stDataFrame"] {
        background-color: #ffffff !important;
        border-radius: 10px !important;
    }

    [data-testid="stFileUploader"] {
        background-color: #ffffff !important;
        color: #17202a !important;
    }

    [data-testid="stFileUploader"] * {
        color: #17202a !important;
    }

    a {
        color: #1f5f8b !important;
    }

    @media (max-width: 768px) {

        .block-container {
            padding-top: 1rem;
            padding-left: 1rem;
            padding-right: 1rem;
        }

        div[data-testid="stVerticalBlock"] > div:has(h1) {
            padding: 25px 20px 22px 20px;
            border-radius: 14px;
        }

        div[data-testid="stVerticalBlock"] h1 {
            font-size: 2.2rem !important;
        }

        div[data-testid="stVerticalBlock"] h1 + p {
            font-size: 1rem !important;
        }

        div.stButton > button[kind="primary"] {
            height: 58px !important;
            font-size: 1.05rem !important;
        }

        div[data-testid="stTextInput"] label,
        div[data-testid="stNumberInput"] label,
        div[data-testid="stDateInput"] label,
        div[data-testid="stSelectbox"] label,
        div[data-testid="stTextArea"] label {
            color: #17202a !important;
            -webkit-text-fill-color: #17202a !important;
            opacity: 1 !important;
            visibility: visible !important;
            display: block !important;
        }

        div[data-testid="stTextInput"] label p,
        div[data-testid="stNumberInput"] label p,
        div[data-testid="stDateInput"] label p,
        div[data-testid="stSelectbox"] label p,
        div[data-testid="stTextArea"] label p {
            color: #17202a !important;
            -webkit-text-fill-color: #17202a !important;
            opacity: 1 !important;
            visibility: visible !important;
            display: block !important;
        }

        div[data-testid="stTextInput"] input,
        div[data-testid="stNumberInput"] input,
        div[data-testid="stDateInput"] input {
            color: #17202a !important;
            -webkit-text-fill-color: #17202a !important;
            background-color: #ffffff !important;
        }
    }

    </style>
    """,
    unsafe_allow_html=True
)


# ============================================================
# AVISO DE PROJETO CALCULADO
# ============================================================

if st.session_state.get("mostrar_sucesso", False):

    st.toast(
        "PROJETO CALCULADO COM SUCESSO!",
        icon="✅"
    )

    st.session_state["mostrar_sucesso"] = False


# ============================================================
# TÍTULO
# ============================================================

st.title("📐 Detalhamento do Projeto")

st.markdown(
    "Preencha as dimensões da estrutura e clique em "
    "**CALCULAR PROJETO** para gerar os quantitativos."
)


# ============================================================
# INFORMAÇÕES DO CLIENTE
# ============================================================

st.header("📋 Informações Gerais")

col_cli1, col_cli2 = st.columns(2)

with col_cli1:

    cliente = st.text_input(
        "Nome do Cliente",
        value="João Silva"
    )

with col_cli2:

    data_projeto = st.date_input(
        "Data do Orçamento",
        date.today()
    )


# ============================================================
# DIMENSÕES
# ============================================================

st.header("🏠 Dimensões da Estrutura")

col_dim1, col_dim2 = st.columns(2)

with col_dim1:

    comprimento_paredes = st.number_input(
        "Comprimento Total das Paredes (m linear)",
        min_value=1.0,
        value=30.0,
        step=1.0
    )

with col_dim2:

    pe_direito = st.number_input(
        "Pé Direito (m)",
        min_value=1.0,
        value=3.0,
        step=0.1
    )


# ============================================================
# ÁREA
# ============================================================

area_total = (
    comprimento_paredes
    * pe_direito
)

st.info(
    f"📐 Área total calculada das paredes: "
    f"**{area_total:.2f} m²**"
)


# ============================================================
# PREÇOS
# ============================================================

PRECOS_BASE = {

    "perfil": 50.0,
    "guia": 50.0,
    "plywood": 80.0,
    "placa_st": 40.0,
    "placa_cimenticia": 140.0,
    "la_pet": 200.0,
    "parafusos": 35.0,
    "massas": 500.0,
    "telas": 500.0,
    "adesivo": 150.0,
    "manta": 1000.0
}


# ============================================================
# QUANTIDADES
# ============================================================

qtd_perfil = area_total * 1.25
qtd_guia = area_total * 0.55
qtd_plywood = area_total / 2.2
qtd_placa_st = area_total / 2.4
qtd_cimenticia = area_total / 2.4
qtd_la = area_total / 10.0
qtd_parafuso = area_total * 0.5
qtd_massa = area_total / 30.0
qtd_tela = area_total / 40.0
qtd_adesivo = area_total / 15.0
qtd_manta = area_total / 50.0


# ============================================================
# LISTA DE MATERIAIS
# ============================================================

lista_materiais = [

    {
        "nome": "Perfil 90x0,80",
        "qtd": qtd_perfil,
        "preco": PRECOS_BASE["perfil"]
    },

    {
        "nome": "Guia Perimetral",
        "qtd": qtd_guia,
        "preco": PRECOS_BASE["guia"]
    },

    {
        "nome": "Plywood 8mm",
        "qtd": qtd_plywood,
        "preco": PRECOS_BASE["plywood"]
    },

    {
        "nome": "Placa ST 12.5mm",
        "qtd": qtd_placa_st,
        "preco": PRECOS_BASE["placa_st"]
    },

    {
        "nome": "Placa Cimentícia 12mm",
        "qtd": qtd_cimenticia,
        "preco": PRECOS_BASE["placa_cimenticia"]
    },

    {
        "nome": "Lã PET",
        "qtd": qtd_la,
        "preco": PRECOS_BASE["la_pet"]
    },

    {
        "nome": "Parafusos (Cento)",
        "qtd": qtd_parafuso,
        "preco": PRECOS_BASE["parafusos"]
    },

    {
        "nome": "Massas (Balde/Saco)",
        "qtd": qtd_massa,
        "preco": PRECOS_BASE["massas"]
    },

    {
        "nome": "Telas (Rolo)",
        "qtd": qtd_tela,
        "preco": PRECOS_BASE["telas"]
    },

    {
        "nome": "Adesivo PU (Cx)",
        "qtd": qtd_adesivo,
        "preco": PRECOS_BASE["adesivo"]
    },

    {
        "nome": "Manta Hidrófuga",
        "qtd": qtd_manta,
        "preco": PRECOS_BASE["manta"]
    },
]


# ============================================================
# CÁLCULO DO PROJETO
# ============================================================

st.markdown("---")

st.subheader("🧮 Cálculo do Projeto")

calcular_projeto = st.button(
    "🧮 CALCULAR PROJETO",
    type="primary",
    use_container_width=True
)


if calcular_projeto:

    # ========================================================
    # NOVAS QUANTIDADES
    # ========================================================

    novas_quantidades = [

        qtd_perfil,
        qtd_guia,
        qtd_plywood,
        qtd_placa_st,
        qtd_cimenticia,
        qtd_la,
        qtd_parafuso,
        qtd_massa,
        qtd_tela,
        qtd_adesivo,
        qtd_manta
    ]


    # ========================================================
    # GRAVA AS 11 QUANTIDADES NO SESSION STATE
    # ========================================================

    for idx, quantidade in enumerate(
        novas_quantidades
    ):

        st.session_state[
            f"q_{idx}"
        ] = float(
            round(
                quantidade,
                1
            )
        )


    # ========================================================
    # MONTA O RESULTADO
    # ========================================================

    dados_calculados = []

    total_materiais_calculado = 0.0


    for idx, mat in enumerate(
        lista_materiais
    ):

        quantidade = float(
            round(
                novas_quantidades[idx],
                1
            )
        )

        preco = float(
            mat["preco"]
        )

        subtotal = (
            quantidade
            * preco
        )

        total_materiais_calculado += subtotal

        dados_calculados.append(
            {
                "Item": mat["nome"],
                "Quantidade": quantidade,
                "Preço Unitário": preco,
                "Total Item": subtotal
            }
        )


    # ========================================================
    # MÃO DE OBRA
    # ========================================================

    mao_de_obra_calculada = 11635.0


    # ========================================================
    # TOTAL GERAL
    # ========================================================

    total_geral_calculado = (

        total_materiais_calculado
        + mao_de_obra_calculada

    )


    # ========================================================
    # GRAVA PROJETO
    # ========================================================

    st.session_state[
        "projeto_calculado"
    ] = {

        "cliente":
            cliente,

        "data_projeto":
            data_projeto,

        "comprimento_paredes":
            comprimento_paredes,

        "pe_direito":
            pe_direito,

        "area_total":
            area_total,

        "dados_atualizados":
            dados_calculados,

        "lista_materiais":
            dados_calculados,

        "total_materiais":
            total_materiais_calculado,

        "mao_de_obra":
            mao_de_obra_calculada,

        "total_geral":
            total_geral_calculado,

        "dimensoes": {

            "comprimento_paredes":
                comprimento_paredes,

            "pe_direito":
                pe_direito,

            "area_paredes":
                area_total
        },

        "calculado":
            True
    }


    # ========================================================
    # MARCA PARA MOSTRAR MENSAGEM APÓS O RERUN
    # ========================================================

    st.session_state[
        "mostrar_sucesso"
    ] = True


    # ========================================================
    # RECARREGA A TELA
    # ========================================================

    st.rerun()


# ============================================================
# INSUMOS
# ============================================================

st.header(
    "📋 Insumos Calculados Automaticamente"
)

st.markdown(
    "As quantidades e os valores unitários podem ser ajustados."
)


dados_atualizados = []

total_materiais = 0.0

col_grid1, col_grid2 = st.columns(2)


for idx, mat in enumerate(
    lista_materiais
):

    coluna_painel = (

        col_grid1
        if idx % 2 == 0
        else col_grid2

    )


    with coluna_painel:

        st.markdown(
            f"""
            <div class="nome-material">
                🔹 {mat["nome"]}
            </div>
            """,
            unsafe_allow_html=True
        )


        c_qtd, c_prc = st.columns(2)


        with c_qtd:

            nova_qtd = st.number_input(

                f"{mat['nome']} (Qtd)",

                min_value=0.0,

                value=float(
                    round(
                        mat["qtd"],
                        1
                    )
                ),

                key=f"q_{idx}"
            )


        with c_prc:

            novo_prc = st.number_input(

                f"{mat['nome']} (Preço R$)",

                min_value=0.0,

                value=float(
                    mat["preco"]
                ),

                key=f"p_{idx}"
            )


        subtotal_calculado = (

            nova_qtd
            * novo_prc

        )


        total_materiais += (
            subtotal_calculado
        )


        dados_atualizados.append(

            {
                "Item":
                    mat["nome"],

                "Quantidade":
                    nova_qtd,

                "Preço Unitário":
                    novo_prc,

                "Total Item":
                    subtotal_calculado
            }

        )


        st.markdown(

            f"""
            <div style="
                background:#eef3f7;
                padding:10px 14px;
                border-radius:8px;
                margin:6px 0 12px 0;
                color:#17202a;
                font-weight:700;
            ">
                Subtotal do Item:
                R$ {subtotal_calculado:,.2f}
            </div>
            """,

            unsafe_allow_html=True
        )


# ============================================================
# MÃO DE OBRA
# ============================================================

mao_de_obra = st.sidebar.number_input(

    "Mão de Obra Geral (R$)",

    min_value=0.0,

    value=11635.0,

    step=100.0
)


# ============================================================
# TOTAL
# ============================================================

total_geral = (

    total_materiais
    + mao_de_obra

)


# ============================================================
# ATUALIZAR CÁLCULO
# ============================================================

st.markdown("---")


if st.button(

    "🔄 ATUALIZAR CÁLCULO",

    use_container_width=True

):

    st.session_state[
        "projeto_calculado"
    ] = {

        "cliente":
            cliente,

        "data_projeto":
            data_projeto,

        "comprimento_paredes":
            comprimento_paredes,

        "pe_direito":
            pe_direito,

        "area_total":
            area_total,

        "dados_atualizados":
            dados_atualizados,

        "lista_materiais":
            dados_atualizados,

        "total_materiais":
            total_materiais,

        "mao_de_obra":
            mao_de_obra,

        "total_geral":
            total_geral,

        "dimensoes": {

            "comprimento_paredes":
                comprimento_paredes,

            "pe_direito":
                pe_direito,

            "area_paredes":
                area_total
        },

        "calculado":
            True
    }

    st.toast(
        "CÁLCULO ATUALIZADO!",
        icon="🔄"
    )


# ============================================================
# RESUMO
# ============================================================

st.header(
    "📊 Resumo Consolidado do Orçamento"
)


df_resumo = pd.DataFrame(
    dados_atualizados
)


st.dataframe(

    df_resumo.style.format(

        {
            "Quantidade":
                "{:.2f}",

            "Preço Unitário":
                "R$ {:.2f}",

            "Total Item":
                "R$ {:.2f}"
        }

    ),

    use_container_width=True
)


# ============================================================
# SIDEBAR
# ============================================================

st.sidebar.header(
    "💰 Custos de Instalação"
)

st.sidebar.markdown("---")


st.sidebar.metric(

    label="Total Materiais",

    value=f"R$ {total_materiais:,.2f}"
)


st.sidebar.metric(

    label="Total Mão de Obra",

    value=f"R$ {mao_de_obra:,.2f}"
)


st.sidebar.subheader(

    f"Total Geral: R$ {total_geral:,.2f}"

)


# ============================================================
# FUNÇÃO — LOCALIZAR LOGO
# ============================================================

def localizar_logo():

    caminhos = [

        Path("assets/logo.png"),
        Path("assets/logo.jpg"),
        Path("assets/logo.jpeg"),
        Path("assets/logo.webp"),

        Path("logo.png"),
        Path("logo.jpg"),
        Path("logo.jpeg"),
        Path("logo.webp"),
    ]


    for caminho in caminhos:

        if caminho.exists():

            return caminho


    return None


# ============================================================
# FUNÇÃO PARA GERAR EXCEL
# ============================================================

def gerar_excel():

    wb = Workbook()

    ws = wb.active

    ws.title = "Orçamento"


    # ========================================================
    # LARGURAS
    # ========================================================

    larguras = {

        "A": 28,
        "B": 18,
        "C": 20,
        "D": 20,
        "E": 22,
        "F": 22,
    }


    for coluna, largura in larguras.items():

        ws.column_dimensions[
            coluna
        ].width = largura


    # ========================================================
    # ESTILOS
    # ========================================================

    fundo_titulo = PatternFill(
        "solid",
        fgColor="17202A"
    )

    fundo_secao = PatternFill(
        "solid",
        fgColor="34495E"
    )

    fundo_cabecalho = PatternFill(
        "solid",
        fgColor="D9E2F3"
    )

    fundo_total = PatternFill(
        "solid",
        fgColor="E2F0D9"
    )

    branco = "FFFFFF"


    fonte_titulo = Font(
        name="Calibri",
        size=20,
        bold=True,
        color=branco
    )

    fonte_secao = Font(
        name="Calibri",
        size=12,
        bold=True,
        color=branco
    )

    fonte_cabecalho = Font(
        name="Calibri",
        size=11,
        bold=True,
        color="17202A"
    )

    fonte_normal = Font(
        name="Calibri",
        size=11,
        color="17202A"
    )

    fonte_total = Font(
        name="Calibri",
        size=13,
        bold=True,
        color="17202A"
    )


    borda_fina = Border(

        left=Side(
            style="thin",
            color="B7B7B7"
        ),

        right=Side(
            style="thin",
            color="B7B7B7"
        ),

        top=Side(
            style="thin",
            color="B7B7B7"
        ),

        bottom=Side(
            style="thin",
            color="B7B7B7"
        ),
    )


    # ========================================================
    # TÍTULO
    # ========================================================

    ws.merge_cells("A1:F2")

    ws["A1"] = "ORÇAMENTO STEEL FRAMING"

    ws["A1"].font = fonte_titulo

    ws["A1"].fill = fundo_titulo

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


    for row in ws["A1:F2"]:

        for cell in row:

            cell.fill = fundo_titulo


    # ========================================================
    # LOGO
    # ========================================================

    ws.merge_cells("A4:B8")


    for row in ws["A4:B8"]:

        for cell in row:

            cell.border = borda_fina

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )


    caminho_logo = localizar_logo()


    if caminho_logo is not None:

        try:

            logo = XLImage(
                str(caminho_logo)
            )

            logo.width = 230
            logo.height = 125
            logo.anchor = "A4"

            ws.add_image(logo)

        except Exception:

            ws["A4"] = (
                "Não foi possível carregar o logo."
            )

            ws["A4"].font = Font(
                size=10,
                color="CC0000"
            )

    else:

        ws["A4"] = (
            "LOGO NÃO ENCONTRADO\n\n"
            "Coloque o arquivo do logo em:\n"
            "assets/logo.png"
        )

        ws["A4"].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        ws["A4"].font = Font(
            size=10,
            bold=True,
            color="666666"
        )

        ws["A4"].fill = PatternFill(
            "solid",
            fgColor="F2F2F2"
        )


    # ========================================================
    # DADOS DA EMPRESA
    # ========================================================

    ws.merge_cells("C4:F4")

    ws["C4"] = "DADOS DA EMPRESA"

    ws["C4"].fill = fundo_secao

    ws["C4"].font = fonte_secao

    ws["C4"].alignment = Alignment(
        horizontal="center"
    )


    dados_empresa = [

        (
            "Empresa",
            "Digite o nome da sua empresa"
        ),

        (
            "CNPJ",
            "Digite o CNPJ"
        ),

        (
            "Telefone / WhatsApp",
            "Digite o telefone"
        ),

        (
            "E-mail",
            "Digite o e-mail"
        ),
    ]


    linha = 5


    for campo, valor in dados_empresa:

        ws[f"C{linha}"] = campo

        ws[f"C{linha}"].font = fonte_cabecalho

        ws.merge_cells(

            start_row=linha,
            start_column=4,
            end_row=linha,
            end_column=6
        )

        ws[f"D{linha}"] = valor

        ws[f"D{linha}"].font = fonte_normal

        linha += 1


    # ========================================================
    # IDENTIFICAÇÃO DO PROJETO
    # ========================================================

    ws.merge_cells("A10:F10")

    ws["A10"] = "IDENTIFICAÇÃO DO PROJETO"

    ws["A10"].fill = fundo_secao

    ws["A10"].font = fonte_secao

    ws["A10"].alignment = Alignment(
        horizontal="center"
    )


    dados_projeto = [

        (
            "Cliente",
            cliente
        ),

        (
            "Data do Orçamento",
            data_projeto.strftime("%d/%m/%Y")
        ),

        (
            "Comprimento das Paredes",
            f"{comprimento_paredes:.2f} m linear"
        ),

        (
            "Pé Direito",
            f"{pe_direito:.2f} m"
        ),

        (
            "Área das Paredes",
            f"{area_total:.2f} m²"
        ),
    ]


    linha = 11


    for campo, valor in dados_projeto:

        ws[f"A{linha}"] = campo

        ws[f"A{linha}"].font = fonte_cabecalho

        ws.merge_cells(

            start_row=linha,
            start_column=2,
            end_row=linha,
            end_column=6
        )

        ws[f"B{linha}"] = valor

        ws[f"B{linha}"].font = fonte_normal

        linha += 1


    # ========================================================
    # MATERIAIS
    # ========================================================

    linha_inicio_materiais = 18


    ws.merge_cells(

        start_row=linha_inicio_materiais,
        start_column=1,
        end_row=linha_inicio_materiais,
        end_column=6
    )


    ws.cell(
        linha_inicio_materiais,
        1
    ).value = "QUANTITATIVO DE MATERIAIS"


    ws.cell(
        linha_inicio_materiais,
        1
    ).fill = fundo_secao


    ws.cell(
        linha_inicio_materiais,
        1
    ).font = fonte_secao


    ws.cell(
        linha_inicio_materiais,
        1
    ).alignment = Alignment(
        horizontal="center"
    )


    # ========================================================
    # CABEÇALHOS
    # ========================================================

    cabecalhos = [

        "Material",
        "Quantidade",
        "Unidade",
        "Preço Unitário",
        "Total",
        "Observação",
    ]


    linha_cabecalho = (
        linha_inicio_materiais + 1
    )


    for col, texto in enumerate(
        cabecalhos,
        start=1
    ):

        cell = ws.cell(
            linha_cabecalho,
            col
        )

        cell.value = texto

        cell.fill = fundo_cabecalho

        cell.font = fonte_cabecalho

        cell.border = borda_fina

        cell.alignment = Alignment(
            horizontal="center"
        )


    # ========================================================
    # ITENS
    # ========================================================

    linha = (
        linha_cabecalho + 1
    )


    for item in dados_atualizados:

        ws.cell(
            linha,
            1
        ).value = item["Item"]


        ws.cell(
            linha,
            2
        ).value = item["Quantidade"]


        ws.cell(
            linha,
            3
        ).value = "un."


        ws.cell(
            linha,
            4
        ).value = item["Preço Unitário"]


        ws.cell(
            linha,
            5
        ).value = (
            f"=B{linha}*D{linha}"
        )


        ws.cell(
            linha,
            6
        ).value = (
            "Quantidade e preço editáveis"
        )


        for col in range(1, 7):

            cell = ws.cell(
                linha,
                col
            )

            cell.border = borda_fina

            cell.font = fonte_normal


        ws.cell(
            linha,
            2
        ).number_format = "0.00"


        ws.cell(
            linha,
            4
        ).number_format = (
            'R$ #,##0.00'
        )


        ws.cell(
            linha,
            5
        ).number_format = (
            'R$ #,##0.00'
        )


        linha += 1


    # ========================================================
    # TOTAL MATERIAIS
    # ========================================================

    linha_total_materiais = (
        linha + 1
    )


    ws.merge_cells(

        start_row=linha_total_materiais,
        start_column=1,
        end_row=linha_total_materiais,
        end_column=4
    )


    ws.cell(
        linha_total_materiais,
        1
    ).value = "TOTAL DE MATERIAIS"


    ws.cell(
        linha_total_materiais,
        1
    ).font = fonte_total


    ws.cell(
        linha_total_materiais,
        1
    ).fill = fundo_total


    primeira_linha = (
        linha_cabecalho + 1
    )


    ultima_linha = (
        linha - 1
    )


    ws.cell(
        linha_total_materiais,
        5
    ).value = (
        f"=SUM(E{primeira_linha}:E{ultima_linha})"
    )


    ws.cell(
        linha_total_materiais,
        5
    ).font = fonte_total


    ws.cell(
        linha_total_materiais,
        5
    ).fill = fundo_total


    ws.cell(
        linha_total_materiais,
        5
    ).number_format = (
        'R$ #,##0.00'
    )


    # ========================================================
    # MÃO DE OBRA
    # ========================================================

    linha_mao_obra = (
        linha_total_materiais + 1
    )


    ws.merge_cells(

        start_row=linha_mao_obra,
        start_column=1,
        end_row=linha_mao_obra,
        end_column=4
    )


    ws.cell(
        linha_mao_obra,
        1
    ).value = "MÃO DE OBRA"


    ws.cell(
        linha_mao_obra,
        1
    ).font = fonte_total


    ws.cell(
        linha_mao_obra,
        1
    ).fill = fundo_total


    ws.cell(
        linha_mao_obra,
        5
    ).value = mao_de_obra


    ws.cell(
        linha_mao_obra,
        5
    ).font = fonte_total


    ws.cell(
        linha_mao_obra,
        5
    ).fill = fundo_total


    ws.cell(
        linha_mao_obra,
        5
    ).number_format = (
        'R$ #,##0.00'
    )


    # ========================================================
    # TOTAL GERAL
    # ========================================================

    linha_total_geral = (
        linha_mao_obra + 1
    )


    ws.merge_cells(

        start_row=linha_total_geral,
        start_column=1,
        end_row=linha_total_geral,
        end_column=4
    )


    ws.cell(
        linha_total_geral,
        1
    ).value = "TOTAL GERAL"


    ws.cell(
        linha_total_geral,
        1
    ).font = Font(
        size=15,
        bold=True,
        color="17202A"
    )


    ws.cell(
        linha_total_geral,
        5
    ).value = (

        f"=E{linha_total_materiais}"
        f"+E{linha_mao_obra}"

    )


    ws.cell(
        linha_total_geral,
        5
    ).font = Font(
        size=15,
        bold=True,
        color="17202A"
    )


    ws.cell(
        linha_total_geral,
        5
    ).number_format = (
        'R$ #,##0.00'
    )


    for col in range(1, 6):

        ws.cell(
            linha_total_geral,
            col
        ).fill = PatternFill(
            "solid",
            fgColor="C6E0B4"
        )


    # ========================================================
    # CONDIÇÕES COMERCIAIS
    # ========================================================

    linha_condicoes = (
        linha_total_geral + 3
    )


    ws.merge_cells(

        start_row=linha_condicoes,
        start_column=1,
        end_row=linha_condicoes,
        end_column=6
    )


    ws.cell(
        linha_condicoes,
        1
    ).value = (
        "CONDIÇÕES COMERCIAIS / OBSERVAÇÕES"
    )


    ws.cell(
        linha_condicoes,
        1
    ).fill = fundo_secao


    ws.cell(
        linha_condicoes,
        1
    ).font = fonte_secao


    for i in range(1, 4):

        linha_obs = (
            linha_condicoes + i
        )


        ws.merge_cells(

            start_row=linha_obs,
            start_column=1,
            end_row=linha_obs,
            end_column=6
        )


        ws.cell(
            linha_obs,
            1
        ).value = (
            "Digite aqui suas condições "
            "comerciais e observações."
        )


        ws.cell(
            linha_obs,
            1
        ).font = fonte_normal


        ws.cell(
            linha_obs,
            1
        ).alignment = Alignment(
            vertical="top",
            wrap_text=True
        )


    # ========================================================
    # CONFIGURAÇÃO DE IMPRESSÃO
    # ========================================================

    ws.freeze_panes = "A20"

    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = "landscape"

    ws.page_setup.paperSize = (
        ws.PAPERSIZE_A4
    )

    ws.page_setup.fitToWidth = 1

    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.print_options.horizontalCentered = True

    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.40
    ws.page_margins.bottom = 0.40
    ws.page_margins.header = 0.20
    ws.page_margins.footer = 0.20

    ws.print_area = (
        f"A1:F{linha_condicoes + 3}"
    )


    # ========================================================
    # ABA MEMÓRIA DE CÁLCULO
    # ========================================================

    memoria = wb.create_sheet(
        "Memória de Cálculo"
    )


    memoria.column_dimensions[
        "A"
    ].width = 35


    memoria.column_dimensions[
        "B"
    ].width = 20


    memoria.column_dimensions[
        "C"
    ].width = 45


    memoria.column_dimensions[
        "D"
    ].width = 20


    memoria.merge_cells(
        "A1:D2"
    )


    memoria["A1"] = (
        "MEMÓRIA DE CÁLCULO"
    )


    memoria["A1"].font = fonte_titulo

    memoria["A1"].fill = fundo_titulo

    memoria["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


    for row in memoria["A1:D2"]:

        for cell in row:

            cell.fill = fundo_titulo


    # ========================================================
    # CABEÇALHO MEMÓRIA
    # ========================================================

    memoria["A4"] = "Parâmetro"
    memoria["B4"] = "Valor"
    memoria["C4"] = "Critério"
    memoria["D4"] = "Unidade"


    for col in range(1, 5):

        cell = memoria.cell(
            4,
            col
        )

        cell.fill = fundo_cabecalho

        cell.font = fonte_cabecalho

        cell.border = borda_fina


    # ========================================================
    # PARÂMETROS
    # ========================================================

    parametros = [

        (
            "Comprimento das Paredes",
            comprimento_paredes,
            "Informado pelo usuário",
            "m linear"
        ),

        (
            "Pé Direito",
            pe_direito,
            "Informado pelo usuário",
            "m"
        ),

        (
            "Área das Paredes",
            area_total,
            "Comprimento × Pé Direito",
            "m²"
        ),

        (
            "Perfil 90x0,80",
            qtd_perfil,
            "Área × 1,25",
            "un."
        ),

        (
            "Guia Perimetral",
            qtd_guia,
            "Área × 0,55",
            "un."
        ),

        (
            "Plywood 8mm",
            qtd_plywood,
            "Área ÷ 2,20",
            "un."
        ),

        (
            "Placa ST 12.5mm",
            qtd_placa_st,
            "Área ÷ 2,40",
            "un."
        ),

        (
            "Placa Cimentícia 12mm",
            qtd_cimenticia,
            "Área ÷ 2,40",
            "un."
        ),

        (
            "Lã PET",
            qtd_la,
            "Área ÷ 10",
            "un."
        ),

        (
            "Parafusos",
            qtd_parafuso,
            "Área × 0,50",
            "centos"
        ),

        (
            "Massas",
            qtd_massa,
            "Área ÷ 30",
            "un."
        ),

        (
            "Telas",
            qtd_tela,
            "Área ÷ 40",
            "rolos"
        ),

        (
            "Adesivo PU",
            qtd_adesivo,
            "Área ÷ 15",
            "caixas"
        ),

        (
            "Manta Hidrófuga",
            qtd_manta,
            "Área das Paredes ÷ 50",
            "un."
        ),
    ]


    linha = 5


    for parametro, valor, criterio, unidade in parametros:

        memoria.cell(
            linha,
            1
        ).value = parametro


        memoria.cell(
            linha,
            2
        ).value = valor


        memoria.cell(
            linha,
            3
        ).value = criterio


        memoria.cell(
            linha,
            4
        ).value = unidade


        for col in range(1, 5):

            memoria.cell(
                linha,
                col
            ).border = borda_fina

            memoria.cell(
                linha,
                col
            ).font = fonte_normal


        linha += 1


    # ========================================================
    # TOTAIS MEMÓRIA
    # ========================================================

    memoria["A22"] = "TOTAL MATERIAIS"

    memoria["B22"] = total_materiais


    memoria["A23"] = "MÃO DE OBRA"

    memoria["B23"] = mao_de_obra


    memoria["A24"] = "TOTAL GERAL"

    memoria["B24"] = total_geral


    for linha_total in [22, 23, 24]:

        memoria.cell(
            linha_total,
            1
        ).font = fonte_total


        memoria.cell(
            linha_total,
            2
        ).font = fonte_total


        memoria.cell(
            linha_total,
            2
        ).number_format = (
            'R$ #,##0.00'
        )


    memoria.sheet_view.showGridLines = False


    # ========================================================
    # ARQUIVO EXCEL
    # ========================================================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output.getvalue()


# ============================================================
# EXPORTAÇÃO
# ============================================================

st.markdown("---")

st.subheader(
    "📥 Exportação do Orçamento"
)


caminho_logo_teste = (
    localizar_logo()
)


if caminho_logo_teste:

    st.success(
        f"✅ Logo encontrado: "
        f"`{caminho_logo_teste}`"
    )

else:

    st.warning(
        "⚠️ Logo não encontrado. "
        "Coloque o arquivo em `assets/logo.png`."
    )


# ============================================================
# GERAR EXCEL
# ============================================================

excel_data = gerar_excel()


# ============================================================
# NOME DO ARQUIVO
# ============================================================

nome_cliente = (

    cliente
    .strip()
    .replace(" ", "_")
    .replace("/", "_")
    .replace("\\", "_")
)


if not nome_cliente:

    nome_cliente = "cliente"


nome_arquivo = (

    f"orcamento_steel_framing_"
    f"{nome_cliente}_"
    f"{data_projeto.strftime('%Y-%m-%d')}.xlsx"

)


# ============================================================
# BOTÃO DOWNLOAD
# ============================================================

st.download_button(

    label=(
        "📊 Baixar Orçamento Profissional "
        "em Excel (.xlsx)"
    ),

    data=excel_data,

    file_name=nome_arquivo,

    mime=(
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    ),

    use_container_width=True,
)


st.caption(
    "O logo é inserido automaticamente na área A4:B8 "
    "do Excel e acompanha a área de impressão."
)
